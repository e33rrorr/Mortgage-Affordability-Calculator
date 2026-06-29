from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
TOTAL_FILL  = PatternFill("solid", fgColor="2E75B6")
ALT_FILL    = PatternFill("solid", fgColor="EBF3FB")

WHITE_BOLD  = Font(name="Calibri", bold=True, color="FFFFFF")
NORMAL_FONT = Font(name="Calibri", color="000000")
DARK_BOLD   = Font(name="Calibri", bold=True, color="1F4E79")

THIN_SIDE   = Side(style="thin", color="BFBFBF")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

CURRENCY_FMT = '£#,##0.00'


class ExpensesExcelExporter:
    """Write a MonthlyExpensesCalculator's data to an Excel file."""

    LABELS = [
        ("Electricity",   "monthly_electricity"),
        ("Water",         "monthly_water"),
        ("Gas",           "monthly_gas"),
        ("Internet",      "monthly_internet"),
        ("Council Tax",   "monthly_council_tax"),
        ("Food",          "monthly_food"),
        ("Transport",     "monthly_transport"),
        ("Loans",         "monthly_loans"),
        ("Credit Cards",  "monthly_credit_cards"),
        ("Childcare",     "monthly_childcare"),
        ("Other",         "monthly_other_expenses"),
    ]

    def __init__(self, expenses_calculator):
        self.calculator = expenses_calculator

    def _style(self, cell, font=None, fill=None, alignment=None,
               number_format=None, border=None):
        if font:          cell.font          = font
        if fill:          cell.fill          = fill
        if alignment:     cell.alignment     = alignment
        if number_format: cell.number_format = number_format
        if border:        cell.border        = border

    def _build_sheet(self, ws):
        ws.title = "Monthly Expenses"
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 18

        # Header row
        for col, text in enumerate(["Category", "Monthly Amount", "Annual Total"], start=1):
            cell = ws.cell(row=1, column=col, value=text)
            self._style(cell, font=WHITE_BOLD, fill=HEADER_FILL,
                        alignment=Alignment(horizontal="center", vertical="center"),
                        border=THIN_BORDER)
        ws.row_dimensions[1].height = 22

        # Data rows
        for row, (label, attr) in enumerate(self.LABELS, start=2):
            monthly_value = getattr(self.calculator, attr, 0)
            fill = ALT_FILL if row % 2 == 0 else None

            label_cell = ws.cell(row=row, column=1, value=label)
            self._style(label_cell, font=NORMAL_FONT, fill=fill,
                        alignment=Alignment(horizontal="left", vertical="center"),
                        border=THIN_BORDER)

            monthly_cell = ws.cell(row=row, column=2, value=monthly_value)
            self._style(monthly_cell, font=NORMAL_FONT, fill=fill,
                        alignment=Alignment(horizontal="center", vertical="center"),
                        number_format=CURRENCY_FMT, border=THIN_BORDER)

            annual_cell = ws.cell(row=row, column=3, value=monthly_value * 12)
            self._style(annual_cell, font=NORMAL_FONT, fill=fill,
                        alignment=Alignment(horizontal="center", vertical="center"),
                        number_format=CURRENCY_FMT, border=THIN_BORDER)

            ws.row_dimensions[row].height = 18

        # Totals row
        total_row = len(self.LABELS) + 2
        first_data_row = 2
        last_data_row  = total_row - 1

        label_cell = ws.cell(row=total_row, column=1, value="Total")
        self._style(label_cell, font=WHITE_BOLD, fill=TOTAL_FILL,
                    alignment=Alignment(horizontal="left", vertical="center"),
                    border=THIN_BORDER)

        for col in (2, 3):
            col_letter = get_column_letter(col)
            formula = f"=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})"
            cell = ws.cell(row=total_row, column=col, value=formula)
            self._style(cell, font=WHITE_BOLD, fill=TOTAL_FILL,
                        alignment=Alignment(horizontal="center", vertical="center"),
                        number_format=CURRENCY_FMT, border=THIN_BORDER)

        ws.row_dimensions[total_row].height = 22
        ws.freeze_panes = "A2"

    def export(self, file_path="monthly_expenses.xlsx"):
        wb = Workbook()
        self._build_sheet(wb.active)
        wb.save(file_path)
        print(f"Expenses saved to {file_path}")
